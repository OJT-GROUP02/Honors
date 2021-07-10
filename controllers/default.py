import datetime
import openpyxl
import psycopg2
import itertools

from operator import itemgetter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, fills, numbers, PatternFill, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.descriptors.excel import UniversalMeasure, Relation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


def index():
    return dict(message="Hello!")


def uro():
    current_date_uro = str(datetime.datetime.today().strftime('%B %d, %Y'))

    header_uro = db((db.header.college_id == 5) |
                    (db.header.semester == '2nd Semester') |
                    (db.header.academic_year == '2019-2020')).select(
        db.college.college_name, db.college.college_address,
        db.header.semester, db.header.academic_year,
        left=db.college.on(db.header.college_id == db.college.college_id))

    male_students = \
        db((db.student.gender == 'M')).select(
            db.student.last_name, db.student.first_name,
            db.student.middle_name, db.student.gender,
            db.grade.total_units, db.grade.sum_of_grades,
            db.grade.final_gwa, db.awards.award_title,
            left=[db.grade.on(db.student.student_id == db.grade.student_id),
                  db.awards.on(db.grade.award_id == db.awards.award_id)],
            orderby=db.student.last_name)

    female_students = \
        db((db.student.gender == 'F')).select(
            db.student.last_name, db.student.first_name,
            db.student.middle_name, db.student.gender,
            db.grade.total_units, db.grade.sum_of_grades,
            db.grade.final_gwa, db.awards.award_title,
            left=[db.grade.on(db.student.student_id == db.grade.student_id),
                  db.awards.on(db.grade.award_id == db.awards.award_id)],
            orderby=db.student.last_name)

    total_units = 234

    return locals()


def rank():
    current_date_rank = str(datetime.datetime.today().strftime('%B %d, %Y'))

    header_rank = db((db.header.college_id == 5) |
                     (db.header.semester == '2nd Semester') |
                     (db.header.academic_year == '2019-2020')).select(
        db.college.college_name, db.college.college_address,
        db.header.semester, db.header.academic_year,
        left=db.college.on(db.header.college_id == db.college.college_id))

    students_honor = db(db.awards.award_title).select(
        db.student.last_name, db.student.first_name,
        db.student.middle_name, db.student.classification,
        db.grade.total_units, db.grade.sum_of_grades,
        db.grade.final_gwa, db.awards.award_title, orderby=db.grade.final_gwa,
        left=[db.grade.on(db.student.student_id == db.grade.student_id),
              db.awards.on(db.grade.award_id == db.awards.award_id)])

    all_students = db(db.student.classification is None).select(
        db.student.last_name, db.student.first_name,
        db.student.middle_name, db.student.classification,
        db.grade.total_units, db.grade.sum_of_grades,
        db.grade.final_gwa, db.awards.award_title,
        left=[db.grade.on(db.student.student_id == db.grade.student_id),
              db.awards.on(db.grade.award_id == db.awards.award_id)])

    return locals()


# db connection
def connect(query):
    cur = psycopg2.connect(database='honorsdb', user='postgres',
                           password='1612', host='localhost',
                            port="5432").cursor()
    cur.execute(query)
    desc = cur.description
    column_names = [col[0] for col in desc]
    result = [dict(itertools.zip_longest(column_names, row)) for row in cur.fetchall()]
    # result = cur.fetchall()

    return result


# QUERIES

# fetch header data
header_query = "SELECT header.header_id, college.college_name, " \
               "college.college_address, header.semester, " \
               "header.academic_year FROM header LEFT JOIN college on " \
               "college.college_id = header.college_id WHERE " \
               "college.college_name = 'College of Business, Economics, " \
                "and Management' \
               AND header.semester = '2nd Semester' AND " \
               "academic_year = '2019-2020'"

header = connect(header_query)
college_name = header[0]["college_name"]
college_name_upper = header[0]["college_name"].upper()
college_address = header[0]["college_address"]
college_semester = header[0]["semester"]
academic_year = header[0]["academic_year"]

# fetch student data
m_student_query = "SELECT CONCAT(student.first_name, ' ', " \
                  "student.middle_name, ' ', student.last_name) AS " \
                  "full_name, grade.total_units, grade.sum_of_grades, " \
                  "grade.final_gwa, awards.award_title FROM student " \
                  "LEFT JOIN grade ON student.student_id = grade.student_id " \
                  "LEFT JOIN awards ON awards.award_id = grade.award_id " \
                  "WHERE student.gender = 'M' ORDER BY student.last_name"

f_student_query = "SELECT CONCAT(student.first_name, ' ', " \
                  "student.middle_name, ' ', student.last_name) AS " \
                  "full_name, grade.total_units, grade.sum_of_grades, " \
                  "grade.final_gwa, awards.award_title FROM student " \
                  "LEFT JOIN grade ON student.student_id = grade.student_id " \
                  "LEFT JOIN awards ON awards.award_id = grade.award_id " \
                  "WHERE student.gender = 'F' ORDER BY student.last_name"

m_students = connect(m_student_query)
m_total_rows = len(m_students)
f_students = connect(f_student_query)
f_total_rows = len(f_students)
minimum_row = 19    # for table
maximum_col = 6     # for table
total_units = 234

# WORKBOOK

wb = Workbook()

wb.create_sheet("Rank", 0)
wb.create_sheet("URO", 1)

ws = wb["URO"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'
ws.page_margins.left = 0.50
ws.page_margins.rigt = 0.50
# ws.print_options.verticalCentered = True
# ws.print_options.horizontalCentered = True

# column sizes
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 9
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 25

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:F1')
ws.append(['Bicol University'])
ws.merge_cells('A2:F2')
ws.append([college_name_upper])
ws.merge_cells('A3:F3')
ws['A3'].font = Font(bold=True)
ws.append([college_address])
ws.merge_cells('A4:F4')

rows = ws.iter_cols(min_row=1, min_col=1, max_row=4, max_col=6)
for row in rows:
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

# Date
date = str(datetime.datetime.today().strftime('%B %d, %Y'))
date_today = ws['F5']
date_today.value = date
date_today.alignment = Alignment(horizontal='right')

# Inside Address
ws['A7'].value = "THE BICOL UNIVERSITY REGISTRAR"
ws['A8'].value = "Bicol University"
ws['A9'].value = "Legazpi City"

for row in ws['A7':'F9']:
    for cell in row:
        cell.alignment = Alignment(horizontal='left')

#Attention
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].bestFit=True

ws['A12'].value = "Attn: UNIVERITY EVALUATION/REVIEW COMMITTEE ON HONOR AND GRADUATES"
ws.merge_cells('A12:F12')
ws['A12'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

#Greetings
ws['A15'].value = "Sir/Madam:"
ws['A15'].alignment = Alignment(horizontal='left')

#Letter Body

for row in range (16, 18):
   ws.row_dimensions[(row)].bestFit=True

ws['A16'].value = "Herewith are the Official List of Candidates for Graduation with Honors under the different degree "
ws['A16'].alignment = Alignment(horizontal='left', indent=2)
ws.merge_cells('A16:F16')
ws['A17'].value = f"programs of the {college_name}  for the {college_semester}, {academic_year}."
ws['A17'].alignment = Alignment(horizontal='left')
ws.merge_cells('A17:F17')

#table

thin = Side(border_style="thin", color="000000")# border style, color
border = Border(left=thin, right=thin, top=thin, bottom=thin)# the position of the border
rows = ws.iter_cols(min_row=19, min_col=1,
                    max_row=minimum_row + m_total_rows + f_total_rows + 3,
                    max_col=6)

for row in rows:
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)   

# table header
ws['A19'].value = "NO."
ws.merge_cells(start_row=19, start_column=1, end_row=20, end_column=1)
ws['B19'].value = "NAME"
ws.merge_cells(start_row=19, start_column=2, end_row=20, end_column=2)
ws['C19'].value = "RATING"
ws.merge_cells('C19:E19')
ws['C20'].value = "Sum of Grades"
ws['D20'].value = "Total Units"
ws['E20'].value = "Final GWA"
ws['F19'].value = "AWARD"
ws.merge_cells(start_row=19, start_column=6, end_row=20, end_column=6)

bold_font = Font(bold=True)
for cell in ws["19:19"]:
    cell.font = bold_font
for cell in ws["20:20"]:
    cell.font = bold_font

# table data font (smaller)
font_size = Font(name='Calibri', size=9)
rows = ws.iter_cols(min_row=21, min_col=1,
                    max_row=minimum_row + m_total_rows + f_total_rows + 3,
                    max_col=6)
for row in rows:
    for cell in row:
        cell.font = font_size

# male students
ws['A21'].value = "MALE"
ws['A21'].font = Font(name='Calibri', size=9, bold=True)
ws.merge_cells('A21:F21')

key_list = ["full_name", "sum_of_grades", "total_units", "final_gwa",
            "award_title"]

min_row_m = minimum_row + 3
max_row_m = min_row_m + m_total_rows
count = 1

temp_row = min_row_m
for m_student in m_students:
    list_index = 0
    ws['A' + str(temp_row)].value = count

    if m_student[key_list[2]] is not None:
        if m_student[key_list[2]] < total_units:
            ws['A' + str(temp_row)].fill = \
                PatternFill(fill_type='solid', fgColor="f0e68c")

    for col_table in range(2, 7):
        char = get_column_letter(col_table)
        ws[char + str(temp_row)] = m_student[key_list[list_index]]

        if m_student[key_list[3]] is not None:
            if list_index == 3:
                ws[char + str(temp_row)].font = \
                    Font(name='Calibri', size=9, bold=True)

        # yellow fill (row)
        if m_student[key_list[2]] is not None:
            if m_student[key_list[2]] < total_units:
                ws[char + str(temp_row)].fill = \
                    PatternFill(fill_type='solid', fgColor="f0e68c")

        # red fill (cell)
        if m_student[key_list[list_index]] is not None:

            if list_index == 3 and m_student[key_list[list_index]] <= 1.75:
                ws[char + str(temp_row)].font = \
                    Font(color="9d0008", name='Calibri', size=9, bold=True)
                ws[char + str(temp_row)].fill = \
                    PatternFill(fill_type='solid', fgColor="fbc3cb")
        list_index += 1

    temp_row += 1
    count += 1

# female students
min_row_f = minimum_row + min_row_m + 2
max_row_f = min_row_f + f_total_rows
count = 1

ws['A' + str(min_row_f - 1)].value = "FEMALE"
ws['A' + str(min_row_f - 1)].font = Font(name='Calibri', size=9, bold=True)
range_f_min = 'A' + str(min_row_f - 1)
range_f_max = 'F' + str(min_row_f - 1)
ws.merge_cells(f'{str(range_f_min)}:{str(range_f_max)}')

temp_row = min_row_f
for f_student in f_students:
    list_index = 0
    ws['A' + str(temp_row)].value = count

    if f_student[key_list[2]] is not None:
        if f_student[key_list[2]] < total_units:
            ws['A' + str(temp_row)].fill = \
                PatternFill(fill_type='solid', fgColor="f0e68c")

    for col_table in range(2, 7):
        char = get_column_letter(col_table)
        ws[char + str(temp_row)] = f_student[key_list[list_index]]

        if f_student[key_list[3]] is not None:
            if list_index == 3:
                ws[char + str(temp_row)].font = \
                    Font(name='Calibri', size=9, bold=True)

        # yellow fill (row)
        if f_student[key_list[2]] is not None:
            if f_student[key_list[2]] < total_units:
                ws[char + str(temp_row)].fill = \
                    PatternFill(fill_type='solid', fgColor="f0e68c")

        # red fill (cell)
        if f_student[key_list[list_index]] is not None:
            if list_index == 3 and f_student[key_list[list_index]] <= 1.75:
                ws[char + str(temp_row)].font = \
                    Font(color="9d0008", name='Calibri', size=9, bold=True)
                ws[char + str(temp_row)].fill = \
                    PatternFill(fill_type='solid', fgColor="fbc3cb")

        list_index += 1

    temp_row += 1
    count += 1

#left align table data
table_total_rows = min_row_m + max_row_f
for table_total_row in range (min_row_m, table_total_rows):
        ws['B' + str(table_total_row)].alignment = Alignment(horizontal='left')

#Contents below the table
total_rows_after_table = minimum_row + m_total_rows + f_total_rows + 5

rows = ws.iter_cols(min_row=19, min_col=1,
                    max_row=minimum_row + m_total_rows + f_total_rows + 3,
                    max_col=6)
        
ws['A' + str(total_rows_after_table)].value = "Note: Subject for verification/recommendation/approval by the University Evaluation/Review"
ws['A' + str(total_rows_after_table)].alignment = Alignment(horizontal='left', indent=2)
range_min = 'A' + str(total_rows_after_table)
range_max = 'F' + str(total_rows_after_table)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['A' + str(total_rows_after_table + 1)].value = "Committee on Honor " \
                                                  "Graduates"
ws['A' + str(total_rows_after_table + 1)].alignment = Alignment(
    horizontal='left', indent=6)
range_min = 'A' + str(total_rows_after_table + 1)
range_max = 'F' + str(total_rows_after_table + 1)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(total_rows_after_table + 3)].value = "Attached are the individual collegiate student's " \
                                                "permanent record with the individual computation of "
ws['A' + str(total_rows_after_table + 3)].alignment = Alignment(
    horizontal='left', indent=2)
range_min = 'A' + str(total_rows_after_table + 3)
range_max = 'F' + str(total_rows_after_table + 3)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(total_rows_after_table + 4)].value = "grades and other important documents of the respective" \
                                                    "honor graduates for comments, recommendation"
range_min = 'A' + str(total_rows_after_table + 4)
range_max = 'F' + str(total_rows_after_table + 4)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['A' + str(total_rows_after_table + 5)].value = " and approval."
range_min = 'A' + str(total_rows_after_table + 5)
range_max = 'F' + str(total_rows_after_table + 5)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(total_rows_after_table + 7)].value = "COLLEGE/CAMPUS EVALUATION/REVIEW COMMITTEE"
ws['A' + str(total_rows_after_table + 7)].alignment = Alignment(
    horizontal='center')
ws['A' + str(total_rows_after_table + 7)].font = Font(bold=True)
range_min = 'A' + str(total_rows_after_table + 7)
range_max = 'F' + str(total_rows_after_table + 7)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

#members
ws['B' + str(total_rows_after_table + 9)].value = "Member N. Name"
ws['B' + str(total_rows_after_table + 9)].alignment = Alignment(
    horizontal='center')
ws['B' + str(total_rows_after_table + 9)].font = Font(
    bold=True)
ws['B' + str(total_rows_after_table + 10)].value = "Member"
ws['B' + str(total_rows_after_table + 10)].alignment = Alignment(
    horizontal='center')
ws['F' + str(total_rows_after_table + 9)].value = "Member N. Name"
ws['F' + str(total_rows_after_table + 9)].alignment = Alignment(
    horizontal='center')
ws['F' + str(total_rows_after_table + 9)].font = Font(
    bold=True)
ws['F' + str(total_rows_after_table + 10)].value = "Member"
ws['F' + str(total_rows_after_table + 10)].alignment = Alignment(
    horizontal='center')

ws['B' + str(total_rows_after_table + 13)].value = "Member N. Name"
ws['B' + str(total_rows_after_table + 13)].alignment = Alignment(
    horizontal='center')
ws['B' + str(total_rows_after_table + 13)].font = Font(
    bold=True)
ws['B' + str(total_rows_after_table + 14)].value = "Member"
ws['B' + str(total_rows_after_table + 14)].alignment = Alignment(
    horizontal='center')
ws['F' + str(total_rows_after_table + 13)].value = "Member N. Name"
ws['F' + str(total_rows_after_table + 13)].alignment = Alignment(
    horizontal='center')
ws['F' + str(total_rows_after_table + 13)].font = Font(
    bold=True)
ws['F' + str(total_rows_after_table + 14)].value = "Member"
ws['F' + str(total_rows_after_table + 14)].alignment = Alignment(
    horizontal='center')

#Co-chair
ws['A' + str(total_rows_after_table + 16)].value = "Co-chair N. Name"
ws['A' + str(total_rows_after_table + 16)].alignment = Alignment(
    horizontal='center')
ws['A' + str(total_rows_after_table + 16)].font = Font(
    bold=True)
range_min = 'A' + str(total_rows_after_table + 16)
range_max = 'F' + str(total_rows_after_table + 16)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(total_rows_after_table + 17)].value = "Co-Chairman"
ws['A' + str(total_rows_after_table + 17)].alignment = Alignment(
    horizontal='center')
range_min = 'A' + str(total_rows_after_table + 17)
range_max = 'F' + str(total_rows_after_table + 17)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

#Chairperson
ws['A' + str(total_rows_after_table + 19)].value = "Chairperson N. Name"
ws['A' + str(total_rows_after_table + 19)].alignment = Alignment(
    horizontal='center')
ws['A' + str(total_rows_after_table + 19)].font = Font(
    bold=True)
range_min = 'A' + str(total_rows_after_table + 19)
range_max = 'F' + str(total_rows_after_table + 19)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(total_rows_after_table + 20)].value = "Chairperson"
ws['A' + str(total_rows_after_table + 20)].alignment = Alignment(
    horizontal='center')
range_min = 'A' + str(total_rows_after_table + 20)
range_max = 'F' + str(total_rows_after_table + 20)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')


#RANK SHEET ---------------------------------------------
# palitan na lang ang arguments

cur = psycopg2.connect(database='honorsdb', user='postgres',
                        password='1612', host='localhost',
                        port="5432").cursor()

ws = wb["Rank"]
ColumnDimension(ws, auto_size=True)
ws.page_setup.paperHeight = '13in'
ws.page_setup.paperWidth = '8.5in'

# column sizes
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 7
ws.column_dimensions["D"].width = 7
ws.column_dimensions["E"].width = 7 
ws.column_dimensions["F"].width = 7
ws.column_dimensions["G"].width = 7
ws.column_dimensions["H"].width = 18

# Heading
ws.append(['Republic of the Philippines'])
ws.merge_cells('A1:H1')
ws.append(['Bicol University'])
ws.merge_cells('A2:H2')
ws.append([college_name_upper])
ws.merge_cells('A3:H3')
ws['A3'].font = Font(bold=True)
ws.append([college_address])
ws.merge_cells('A4:H4')

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(
            wrap_text=True, horizontal='center', vertical='center')
        

# Date
date = str(datetime.datetime.today().strftime('%B %d, %Y'))
date_today = ws['H5']
date_today.value = date
date_today.alignment = Alignment(horizontal='right')


# Inside Address
ws['A7'].value = "THE BICOL UNIVERSITY REGISTRAR"
ws['A8'].value = "Bicol University"
ws['A9'].value = "Legazpi City"

for row in ws['A7':'H9']:
    for cell in row:
        cell.alignment = Alignment(horizontal='left')
        
#Attention   
for col in range(1, 7):
    ws.column_dimensions[get_column_letter(col)].bestFit=True

ws['A12'].value = "Attn: UNIVERITY EVALUATION/REVIEW COMMITTEE ON HONOR AND GRADUATES"
ws.merge_cells('A12:H12')
ws['A12'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

#Greetings
ws['A15'].value = "Sir/Madam:"
ws['A15'].alignment = Alignment(horizontal='left')

#Letter Body

for row in range (16, 18):
   ws.row_dimensions[(row)].bestFit=True
   
ws['A16'].value = "Herewith are the Official List of Candidates for Graduation with Honors under the different degree programs"
ws['A16'].alignment = Alignment(horizontal='left', indent=1)
ws.merge_cells('A16:H16')

ws['A17'].value = f" of the {college_name}  for the {college_semester}, {academic_year} "
ws['A17'].alignment = Alignment(horizontal='left')
ws.merge_cells('A17:H17')

#table
thin = Side(border_style="thin", color="000000")# border style, color 
border = Border(left=thin, right=thin, top=thin, bottom=thin)# the position of the border 

rows = ws.iter_cols(min_row=19, min_col=1, max_row=77, max_col=8)

for row in rows:
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

ws['A19'].value = "NO"
ws.merge_cells(start_row=19, start_column=1, end_row=20, end_column=1)
ws['B19'].value = "NAME"
ws.merge_cells(start_row=19, start_column=2, end_row=20, end_column=2)
ws['C19'].value = "CLASSIFICATION"
ws.merge_cells(start_row=19, start_column=3, end_row=19, end_column=4)
ws['C20'].value = "Regular"
ws['D20'].value = "Irregular"
ws['E19'].value = "RATING"
ws.merge_cells(start_row=19, start_column=5, end_row=19, end_column=7)
ws['E20'].value = "Total Units"
ws['F20'].value = "Sum of Grades"
ws['G20'].value = "Final GWA"
ws['H19'].value = "AWARDS"
ws.merge_cells(start_row=19, start_column=8, end_row=20, end_column=8)

for cell in ws["19:19"]:
    cell.font = font_size = Font(name='Calibri', size=10, bold=True)
for cell in ws["20:20"]:
    cell.font = font_size = Font(name='Calibri', size=10, bold=True)

font_size = Font(name='Calibri', size=9)
rows = ws.iter_cols(min_row=21, min_col=1,
                    max_row=minimum_row + m_total_rows + f_total_rows + 3,
                    max_col=8)
for row in rows:
    for cell in row:
        cell.font = font_size


# ----------------------- TABLE 1 -----------------------
# Table data
student_list = []
cur.execute('SELECT * from student')
for student in cur:
    stud_row = list(student)
    student_list.append(stud_row)

grade_list = []
cur.execute('SELECT * from grade')
for grade in cur:
    grade_row = list(grade)
    grade_list.append(grade_row)
            
award_list = []
cur.execute('SELECT * from awards')
for award in cur:
    award_row = list(award)
    award_list.append(award_row)

student_list.sort()
grade_list.sort()
    
table_1_num = 0
total_stud_num = len(student_list)
for i in range(0, len(student_list)):
    if student_list[i][5] != None:
        table_1_num += 1                                         # total rows ng table 1

start_table_1_num_row = 21                                       # row start ng table 1
end_table_1_num_row = table_1_num + start_table_1_num_row        # row end ng table 1

for i in range(0,table_1_num + 1):

    # Grade items
    grade_units = grade_list[i][2]
    grade_id = grade_list[i][5]         
    grade_sum = grade_list[i][3]
    grade_gwa = grade_list[i][4]

    # Student items
    stud_name = student_list[i][2] + ' ' + student_list[i][3] + ' ' + student_list[i][1]
    stud_classification = student_list[i][5]

    # Award items
    if grade_id != None:
        award_id = award_list[grade_id - 1][0]
        award_title = award_list[grade_id - 1][1]

        reg_tuple = (str(i), stud_name, '✔','', grade_units, grade_sum, grade_gwa, award_title)
        irreg_tuple = (str(i), stud_name, '','✔', grade_units, grade_sum, grade_gwa, award_title)

        reg_list = list(reg_tuple)
        irreg_list = list(irreg_tuple)

        if stud_classification == 'Regular':
            for j in range(1, len(reg_list) + 1):
                char = get_column_letter(j)
                if reg_list[j - 1] == grade_gwa:
                    ws[char + str(start_table_1_num_row)].value = reg_list[j - 1]
                    ws[char + str(start_table_1_num_row)].font = \
                        Font(color="9d0008", name='Calibri', size=9, bold=True)
                    ws[char + str(start_table_1_num_row)].fill = \
                        PatternFill(fill_type='solid', fgColor="fbc3cb")
                else:
                    ws[char + str(start_table_1_num_row)].value = reg_list[j - 1]

        if stud_classification == 'Irregular':
            for j in range(1, len(irreg_list) + 1):
                char = get_column_letter(j)
                if irreg_list[j - 1] == grade_gwa:
                    ws[char + str(start_table_1_num_row)].value = irreg_list[j - 1]
                    ws[char + str(start_table_1_num_row)].font = \
                        Font(color="9d0008", name='Calibri', size=9, bold=True)
                    ws[char + str(start_table_1_num_row)].fill = \
                        PatternFill(fill_type='solid', fgColor="fbc3cb")
                else:
                    ws[char + str(start_table_1_num_row)].value = irreg_list[j - 1]
        
        start_table_1_num_row += 1

# ----------------------- END OF TABLE 1 -----------------------
start_row_committee = end_table_1_num_row + 1
# start_row_committee                                                   -> row(81) start para sa data ng committee
# ws['A' + str(start_row_committee)].value = "EVALUATION COMMITTEE:"    -> example use
        
#Committee
ws['A' + str(start_row_committee)].value = "EVALUATION COMMITTEE:"
ws['A' + str(start_row_committee)].font = Font(bold=True)
range_min = 'A' + str(start_row_committee)
range_max = 'F' + str(start_row_committee)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['B' + str(start_row_committee + 2)].value = "_________________________"
ws['B' + str(start_row_committee + 2)].font = Font(bold=True)
ws['B' + str(start_row_committee + 2)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'B' + str(start_row_committee + 2)
range_max = 'C' + str(start_row_committee + 2)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['B' + str(start_row_committee + 3)].value = "Committee"
ws['B' + str(start_row_committee + 3)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'B' + str(start_row_committee + 3)
range_max = 'C' + str(start_row_committee + 3)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['F' + str(start_row_committee + 2)].value = '_________________________'
ws['F' + str(start_row_committee + 2)].font = Font(bold=True)
range_min = 'F' + str(start_row_committee + 2)
range_max = 'G' + str(start_row_committee + 2)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['F' + str(start_row_committee + 3)].value = "Committee"
ws['F' + str(start_row_committee + 3)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'F' + str(start_row_committee + 3)
range_max = 'G' + str(start_row_committee + 3)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['B' + str(start_row_committee + 7)].value = '_________________________'
ws['B' + str(start_row_committee + 7)].font = Font(bold=True)
ws['B' + str(start_row_committee + 7)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'B' + str(start_row_committee + 7)
range_max = 'C' + str(start_row_committee + 7)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['F' + str(start_row_committee + 7)].value = '_________________________'
ws['F' + str(start_row_committee + 2)].font = Font(bold=True)
range_min = 'F' + str(start_row_committee + 7)
range_max = 'G' + str(start_row_committee + 7)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['B' + str(start_row_committee + 8)].value = "Committee"
ws['B' + str(start_row_committee + 8)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'B' + str(start_row_committee + 8)
range_max = 'C' + str(start_row_committee + 8)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['F' + str(start_row_committee + 8)].value = "Committee"
ws['F' + str(start_row_committee + 8)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'F' + str(start_row_committee + 8)
range_max = 'G' + str(start_row_committee + 8)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['C' + str(start_row_committee + 13)].value = "Noted:"
ws['C' + str(start_row_committee + 13)].font = Font(size=11)

ws['D' + str(start_row_committee + 13)].value = '_________________________'
ws['D' + str(start_row_committee + 13)].font = Font(bold=True)
ws['D' + str(start_row_committee + 13)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'D' + str(start_row_committee + 13)
range_max = 'F' + str(start_row_committee + 13)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['D' + str(start_row_committee + 14)].value = "Position"
ws['D' + str(start_row_committee + 14)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'D' + str(start_row_committee + 14)
range_max = 'F' + str(start_row_committee + 14)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')


# #Note
ws['B' + str(start_row_committee + 25)].value = "Note: Subject for verification/recommendation/approval by the University Evaluation/Review Committee"
ws['B' + str(start_row_committee + 25)].alignment = Alignment(horizontal='left', indent=1, wrapText=True)
range_min = 'B' + str(start_row_committee + 25)
range_max = 'H' + str(start_row_committee + 25)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['B' + str(start_row_committee + 26)].value = "on Honor Graduates"
ws['B' + str(start_row_committee + 26)].alignment = Alignment(horizontal='left', indent=1, wrapText=True)
range_min = 'B' + str(start_row_committee + 26)
range_max = 'C' + str(start_row_committee + 26)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(start_row_committee + 28)].value = "Attached are the individual collegiate student's permanent record with the individual computation of grades"
ws['A' + str(start_row_committee + 28)].alignment = Alignment(horizontal='left', indent=1, wrapText=True)
range_min = 'A' + str(start_row_committee + 28)
range_max = 'H' + str(start_row_committee + 28)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(start_row_committee + 29)].value = "and other important documents of the respective honor graduates for comments, recommendations"
ws['A' + str(start_row_committee + 29)].alignment = Alignment(horizontal='left', wrapText=True)
ws['A' + str(start_row_committee + 29)].font = Font(size=9)
range_min = 'A' + str(start_row_committee + 29)
range_max = 'H' + str(start_row_committee + 29)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(start_row_committee + 30)].value = "and approval"
ws['A' + str(start_row_committee + 30)].alignment = Alignment(horizontal='left', wrapText=True)
ws['A' + str(start_row_committee + 30)].font = Font(size=9)
range_min = 'A' + str(start_row_committee + 30)
range_max = 'H' + str(start_row_committee + 30)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(start_row_committee + 35)].value = "COLLEGE/CAMPUS EVALUATION/REVIEW COMMITTEE"
ws['A' + str(start_row_committee + 35)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['A' + str(start_row_committee + 35)].font = Font(bold=True)
range_min = 'A' + str(start_row_committee + 35)
range_max = 'H' + str(start_row_committee + 35)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['A' + str(start_row_committee + 37)].value = '_________________________'
ws['A' + str(start_row_committee + 37)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['A' + str(start_row_committee + 37)].font = Font(bold=True)
range_min = 'A' + str(start_row_committee + 37)
range_max = 'B' + str(start_row_committee + 37)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['A' + str(start_row_committee + 38)].value = 'Member'
ws['A' + str(start_row_committee + 38)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'A' + str(start_row_committee + 38)
range_max = 'B' + str(start_row_committee + 38)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['C' + str(start_row_committee + 37)].value = '_________________________'
ws['C' + str(start_row_committee + 37)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['C' + str(start_row_committee + 37)].font = Font(bold=True)
range_min = 'C' + str(start_row_committee + 37)
range_max = 'D' + str(start_row_committee + 37)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['C' + str(start_row_committee + 38)].value = 'Member'
ws['C' + str(start_row_committee + 38)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'C' + str(start_row_committee + 38)
range_max = 'D' + str(start_row_committee + 38)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['E' + str(start_row_committee + 37)].value = '_________________________'
ws['E' + str(start_row_committee + 37)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['E' + str(start_row_committee + 37)].font = Font(bold=True)
range_min = 'E' + str(start_row_committee + 37)
range_max = 'F' + str(start_row_committee + 37)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')
ws['E' + str(start_row_committee + 38)].value = 'Member'
ws['E' + str(start_row_committee + 38)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'E' + str(start_row_committee + 38)
range_max = 'F' + str(start_row_committee + 38)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['G' + str(start_row_committee + 37)].value = '_________________________'
ws['G' + str(start_row_committee + 37)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['G' + str(start_row_committee + 37)].font = Font(bold=True)
range_min = 'G' + str(start_row_committee + 37)
range_max = 'H' + str(start_row_committee + 37)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['G' + str(start_row_committee + 38)].value = 'Member'
ws['G' + str(start_row_committee + 38)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'G' + str(start_row_committee + 38)
range_max = 'H' + str(start_row_committee + 38)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['D' + str(start_row_committee + 41)].value = '_________________________'
ws['D' + str(start_row_committee + 41)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['D' + str(start_row_committee + 41)].font = Font(bold=True)
range_min = 'D' + str(start_row_committee + 41)
range_max = 'E' + str(start_row_committee + 41)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['D' + str(start_row_committee + 42)].value = "Co-Chairman"
ws['D' + str(start_row_committee + 42)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'D' + str(start_row_committee + 42)
range_max = 'E' + str(start_row_committee + 42)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['D' + str(start_row_committee + 44)].value = '_________________________'
ws['D' + str(start_row_committee + 44)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
ws['D' + str(start_row_committee + 44)].font = Font(bold=True)
range_min = 'D' + str(start_row_committee + 44)
range_max = 'E' + str(start_row_committee + 44)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

ws['D' + str(start_row_committee + 45)].value = "Chairperson"
ws['D' + str(start_row_committee + 45)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
range_min = 'D' + str(start_row_committee + 45)
range_max = 'E' + str(start_row_committee + 45)
ws.merge_cells(f'{str(range_min)}:{str(range_max)}')

# ----------------------- TABLE 2 -----------------------
honor_stud = []
other_stud = []

# Sorting of student according to honor roles
for i in range(0, len(student_list)):
    if student_list[i][5] != None:
        honor_stud.append(student_list[i])
    else:
        other_stud.append(student_list[i])
    
i = 57  
j = 1

# setting the starting row for table 2
table_xy = coordinate_from_string('A' + str(start_row_committee + 49)) # returns ('A',4)
start_table_2_num_row = table_xy[1] #access index 1 which is the row num

max_row_other_stud =  start_table_2_num_row + len(other_stud) - 1

thin = Side(border_style="thin", color="000000")# border style, color 
border = Border(left=thin, right=thin, top=thin, bottom=thin)# the position of the border 

rows = ws.iter_cols(min_row=start_table_2_num_row, min_col=1, max_row=max_row_other_stud, max_col=8)
for row in rows:
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        cell.font = Font(size=9)

# table 2 loop
for stud in other_stud:

    # Student items
    stud_id = stud[0]
    stud_name = stud[2] + ' ' + stud[3] + ' ' + stud[1]
                
    # Grade items
    grade_units = grade_list[i][2]
    grade_id = grade_list[i][5]         
    grade_sum = grade_list[i][3]
    grade_gwa = grade_list[i][4]

    none_tuple = (str(i + 1), stud_name, '','', grade_units, grade_sum, grade_gwa)
    none_list = list(none_tuple)

    if stud[2] != 'Jhon Christian':
        if grade_units == 234:
            for k in range(1, len(none_list) + 1):
                char = get_column_letter(k)
                ws[char + str(start_table_2_num_row)].value = none_list[k - 1]
            i += 1  
            start_table_2_num_row += 1 
            
        # yellow 
        else:
            if grade_units != None:
                for k in range(1, len(none_list) + 1):
                    char = get_column_letter(k)
                    ws[char + str(start_table_2_num_row)].value = none_list[k - 1]
                    ws[char + str(start_table_2_num_row)].font = \
                        Font(color="9d0008", name='Calibri', size=9, bold=True)
                    ws[char + str(start_table_2_num_row)].fill = \
                        PatternFill(fill_type='solid', fgColor="f0e68c")
                    if k == 7:
                        ws[char + str(start_table_2_num_row)].fill = \
                        PatternFill(fill_type='solid', fgColor="fbc3cb") 
            else:
                for k in range(1, len(none_list) + 1):
                    char = get_column_letter(k)
                    ws[char + str(start_table_2_num_row)].value = none_list[k - 1]
            i += 1  
            start_table_2_num_row += 1 
         
    if j == 4:
        for stud in other_stud:
            if stud[2] == 'Jhon Christian':
                stud_name = stud[2] + ' ' + stud[3] + ' ' + stud[1]
                grade_units = grade_list[0][2]
                grade_sum = grade_list[0][3]
                grade_gwa = grade_list[0][4]
                for k in range(1, len(none_list) + 1):
                    char = get_column_letter(k)
                    none_tuple = (str(i + 1), stud_name, '','', grade_units, grade_sum, grade_gwa)
                    none_list = list(none_tuple)
                    ws[char + str(start_table_2_num_row)].value = none_list[k - 1]  
        i += 1
        start_table_2_num_row += 1
    j += 1
    

# ----------------------- END OF TABLE 2 -----------------------

wb.save('static/Honors.xlsx')
# ivee
#wb.save('D:\\Users\\iveej\\Desktop\\web2py\\applications\\honors\\static'
#        '\\Honors.xlsx')
